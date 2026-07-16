# backend/src/services/excel_writer.py
"""
Excel writer for AXIA statement extraction output.

Replicates the exact column layout of 4751R.xlsx:
  A: TRADE DATE       B: DELIVERY / PRODUCT   C: LONG    D: SHORT
  E: REALIZED PnL     F: COMMISION FEES        G: MARKET FEES
  H: NFA FEES         I: TOTAL COMMS (formula) J: TOTAL PnL (formula)
  K: CURRENCY
"""

from io import BytesIO
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_HEADER_FILL   = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)

_ROW_FILL_ODD  = PatternFill("solid", fgColor="F2F2F2")
_ROW_FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")
_ROW_FONT      = Font(name="Calibri", size=10)

_TOTAL_FILL    = PatternFill("solid", fgColor="D9E1F2")
_TOTAL_FONT    = Font(name="Calibri", bold=True, size=10)

_THIN   = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FMT_DATE  = "DD-MMM-YYYY"
_FMT_INT   = "#,##0"
_FMT_NUM2  = "#,##0.00;[RED]-#,##0.00"


# ---------------------------------------------------------------------------
# Column definitions
# ---------------------------------------------------------------------------

COLUMNS = [
    ("TRADE DATE",          14, "center", _FMT_DATE),
    ("DELIVERY / PRODUCT",  26, "left",   "@"),
    ("LONG",                 9, "center", _FMT_INT),
    ("SHORT",                9, "center", _FMT_INT),
    ("REALIZED PnL",        14, "right",  _FMT_NUM2),
    ("COMMISION FEES",      14, "right",  _FMT_NUM2),
    ("MARKET FEES",         12, "right",  _FMT_NUM2),
    ("NFA FEES",            11, "right",  _FMT_NUM2),
    ("TOTAL COMMS",         12, "right",  _FMT_NUM2),
    ("TOTAL PnL",           12, "right",  _FMT_NUM2),
    ("CURRENCY",            10, "center", "@"),
]

_HDR = {col[0]: idx + 1 for idx, col in enumerate(COLUMNS)}


# ---------------------------------------------------------------------------
# Writer
# ---------------------------------------------------------------------------

def write_statement_excel(
    data: dict,
    output_path: Optional[str | Path] = None,
) -> bytes:
    """
    Write extracted statement data to an Excel file matching the 4751R.xlsx layout.

    Args:
        data:        Output of statement_service.extract_statement()
        output_path: Optional path to save. If None, returns bytes only.

    Returns:
        Raw Excel bytes (always).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statement"
    ws.freeze_panes = "A2"

    # ---- Header row ----
    for col_idx, (header, width, align, fmt) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN
        cell.border    = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30

    # ---- Data rows ----
    rows = data.get("rows", [])
    for row_idx, row in enumerate(rows, start=2):
        fill = _ROW_FILL_ODD if row_idx % 2 != 0 else _ROW_FILL_EVEN
        col  = _HDR

        def _cell(c, val, align="right", fmt=None):
            cell = ws.cell(row=row_idx, column=c, value=val)
            cell.font      = _ROW_FONT
            cell.fill      = fill
            cell.border    = _BORDER
            cell.alignment = Alignment(horizontal=align, vertical="center")
            if fmt:
                cell.number_format = fmt
            return cell

        from datetime import datetime
        td_raw = row.get("trade_date", "")
        try:
            td_val = datetime.strptime(td_raw, "%Y-%m-%d")
        except (ValueError, TypeError):
            td_val = td_raw

        _cell(col["TRADE DATE"],         td_val,                         "center", _FMT_DATE)
        _cell(col["DELIVERY / PRODUCT"], row.get("delivery_product", ""), "left",   "@")
        _cell(col["LONG"],               row.get("long"),                 "center", _FMT_INT)
        _cell(col["SHORT"],              row.get("short"),                "center", _FMT_INT)
        _cell(col["REALIZED PnL"],       row.get("realized_pnl"),        "right",  _FMT_NUM2)
        _cell(col["COMMISION FEES"],     row.get("commission_fees"),     "right",  _FMT_NUM2)
        _cell(col["MARKET FEES"],        row.get("market_fees"),         "right",  _FMT_NUM2)
        _cell(col["NFA FEES"],           row.get("nfa_fees"),            "right",  _FMT_NUM2)

        f_col = get_column_letter(col["COMMISION FEES"])
        g_col = get_column_letter(col["MARKET FEES"])
        h_col = get_column_letter(col["NFA FEES"])
        comm_cell = ws.cell(
            row=row_idx, column=col["TOTAL COMMS"],
            value=f"={f_col}{row_idx}+{g_col}{row_idx}+{h_col}{row_idx}",
        )
        comm_cell.font          = _ROW_FONT
        comm_cell.fill          = fill
        comm_cell.border        = _BORDER
        comm_cell.alignment     = Alignment(horizontal="right", vertical="center")
        comm_cell.number_format = _FMT_NUM2

        e_col = get_column_letter(col["REALIZED PnL"])
        i_col = get_column_letter(col["TOTAL COMMS"])
        pnl_cell = ws.cell(
            row=row_idx, column=col["TOTAL PnL"],
            value=f"={e_col}{row_idx}+{i_col}{row_idx}",
        )
        pnl_cell.font          = _ROW_FONT
        pnl_cell.fill          = fill
        pnl_cell.border        = _BORDER
        pnl_cell.alignment     = Alignment(horizontal="right", vertical="center")
        pnl_cell.number_format = _FMT_NUM2

        _cell(col["CURRENCY"], row.get("currency", ""), "center", "@")

    # ---- Totals row (per currency) ----
    if rows:
        last_data_row = len(rows) + 1
        totals_row    = last_data_row + 2

        ws.cell(row=totals_row, column=1, value="TOTALS").font = _TOTAL_FONT

        from collections import defaultdict
        ccy_rows: dict = defaultdict(list)
        for r_idx, row in enumerate(rows, start=2):
            ccy_rows[row.get("currency", "")].append(r_idx)

        offset = 0
        for ccy, r_idxs in ccy_rows.items():
            t_row    = totals_row + offset
            ccy_cell = ws.cell(row=t_row, column=_HDR["CURRENCY"], value=ccy)
            ccy_cell.font      = _TOTAL_FONT
            ccy_cell.fill      = _TOTAL_FILL
            ccy_cell.border    = _BORDER
            ccy_cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_name in ("LONG", "SHORT", "REALIZED PnL",
                             "COMMISION FEES", "MARKET FEES", "NFA FEES",
                             "TOTAL COMMS", "TOTAL PnL"):
                col_idx  = _HDR[col_name]
                col_ltr  = get_column_letter(col_idx)
                refs     = "+".join(f"{col_ltr}{r}" for r in r_idxs)
                fmt      = _FMT_INT if col_name in ("LONG", "SHORT") else _FMT_NUM2
                tc = ws.cell(row=t_row, column=col_idx, value=f"={refs}")
                tc.font          = _TOTAL_FONT
                tc.fill          = _TOTAL_FILL
                tc.border        = _BORDER
                tc.number_format = fmt
                tc.alignment     = Alignment(horizontal="right", vertical="center")

            offset += 1

    # ---- Metadata sheet ----
    meta      = wb.create_sheet("Info")
    meta["A1"] = "Trade Date"
    meta["B1"] = data.get("trade_date", "")
    meta["A2"] = "Client"
    meta["B2"] = data.get("client", "")
    meta["A3"] = "Account"
    meta["B3"] = data.get("account", "")
    meta["A4"] = "Rows Extracted"
    meta["B4"] = len(rows)

    # ---- Save ----
    buf = BytesIO()
    wb.save(buf)
    raw = buf.getvalue()

    if output_path:
        Path(output_path).write_bytes(raw)

    return raw
