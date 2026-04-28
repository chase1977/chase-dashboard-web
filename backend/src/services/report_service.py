# backend/src/services/report_service.py
"""
Generates downloadable reports:
  - Excel (.xlsx)  — full institutional metrics workbook
  - PDF            — investor-facing summary (via ReportLab)
  - CSV            — raw snapshot data

All functions return bytes so FastAPI can stream them directly.
"""

import io
import pandas as pd
from datetime import datetime
from typing import Optional

from src.services import data_service as ds

# ---------------------------------------------------------------------------
# Config — colours used in the Excel workbook
# ---------------------------------------------------------------------------

# Dark institutional palette (ARGB hex, no leading #)
EXCEL_BG_DARK    = "FF0B1420"   # page background
EXCEL_BG_HEADER  = "FF0D1B2E"   # table header
EXCEL_BG_ROW_ALT = "FF111C2B"   # alternating row
EXCEL_BG_ROW     = "FF0D1728"   # base row
EXCEL_ACCENT     = "FF0EA5E9"   # teal-blue accent
EXCEL_POS        = "FF34D399"   # positive return (green)
EXCEL_NEG        = "FFF87171"   # negative return (red)
EXCEL_TEXT       = "FFE2E8F0"   # primary text
EXCEL_TEXT_MUTED = "FF64748B"   # muted text
EXCEL_BORDER     = "FF1E3A5F"   # cell border


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def generate_excel(data_dir: str, entity_id: str = "portfolio_main") -> bytes:
    """
    Build a multi-sheet Excel workbook for the given entity.
    Sheets: Summary | Pods | Strategies | Traders | Equity Curve
    """
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                 numbers as xl_numbers)
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    snaps    = ds.get_snapshots(data_dir)
    entities = ds.get_entities(data_dir)
    snap_map = snaps.sort_values("timestamp").groupby("entity_id").last()

    # --- Helper styles ---
    def hdr_fill():   return PatternFill("solid", fgColor=EXCEL_BG_HEADER)
    def row_fill(i):  return PatternFill("solid", fgColor=EXCEL_BG_ROW_ALT if i % 2 else EXCEL_BG_ROW)
    def accent_fill():return PatternFill("solid", fgColor=EXCEL_ACCENT)
    def hdr_font():   return Font(name="Calibri", bold=True,  color=EXCEL_TEXT,       size=10)
    def body_font():  return Font(name="Calibri", bold=False, color=EXCEL_TEXT,       size=10)
    def title_font(): return Font(name="Calibri", bold=True,  color=EXCEL_ACCENT,     size=13)
    def meta_font():  return Font(name="Calibri", bold=False, color=EXCEL_TEXT_MUTED, size=9)
    def thin_border():
        s = Side(style="thin", color=EXCEL_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)
    def centre():     return Alignment(horizontal="center", vertical="center")
    def right():      return Alignment(horizontal="right",  vertical="center")
    def left():       return Alignment(horizontal="left",   vertical="center", wrap_text=False)

    # --- Sheet builder helper ---
    def add_sheet(name: str, col_headers: list, data_rows: list,
                  col_widths: list, title: str, subtitle: str):
        ws = wb.create_sheet(name)
        ws.sheet_view.showGridLines = False

        # Title block (rows 1-3)
        ws.merge_cells("A1:J1")
        c = ws["A1"]
        c.value     = title
        c.font      = title_font()
        c.alignment = left()
        c.fill      = PatternFill("solid", fgColor=EXCEL_BG_DARK)

        ws.merge_cells("A2:J2")
        c = ws["A2"]
        c.value     = subtitle
        c.font      = meta_font()
        c.alignment = left()
        c.fill      = PatternFill("solid", fgColor=EXCEL_BG_DARK)

        ws.merge_cells("A3:J3")
        ws["A3"].fill = PatternFill("solid", fgColor=EXCEL_BG_DARK)
        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 14
        ws.row_dimensions[3].height = 6

        # Header row (row 4)
        for ci, h in enumerate(col_headers, 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font      = hdr_font()
            c.fill      = hdr_fill()
            c.alignment = centre() if ci > 1 else left()
            c.border    = thin_border()
        ws.row_dimensions[4].height = 18

        # Data rows
        for ri, row_data in enumerate(data_rows, 5):
            fill = row_fill(ri)
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill   = fill
                c.border = thin_border()
                c.font   = body_font()
                # Colour-code pct columns (cols 6-8 by convention)
                if ci in (6, 7, 8) and isinstance(val, (int, float)):
                    if val > 0:
                        c.font = Font(name="Calibri", color=EXCEL_POS, size=10)
                    elif val < 0:
                        c.font = Font(name="Calibri", color=EXCEL_NEG, size=10)
                c.alignment = centre() if ci > 1 else left()
            ws.row_dimensions[ri].height = 16

        # Column widths
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

        # Freeze panes below header
        ws.freeze_panes = "A5"
        return ws

    # --- Sheet 1: Portfolio Summary ---
    port_snap = snap_map.loc["portfolio_main"] if "portfolio_main" in snap_map.index else None
    if port_snap is not None:
        kpis = ds.snapshot_kpis(port_snap)
        ws_summary = wb.create_sheet("Summary")
        ws_summary.sheet_view.showGridLines = False
        _write_summary_sheet(ws_summary, kpis, title_font, meta_font, body_font,
                             hdr_fill, hdr_font, thin_border, left, right, centre,
                             PatternFill, Font, Alignment, EXCEL_BG_DARK,
                             EXCEL_BG_ROW, EXCEL_ACCENT, EXCEL_TEXT, EXCEL_TEXT_MUTED)

    # --- Helper: build breakdown rows for an entity_type ---
    def _type_rows(etype: str):
        subset = entities[entities["entity_type"] == etype]
        rows = []
        for _, ent in subset.iterrows():
            if ent["entity_id"] not in snap_map.index:
                continue
            s = snap_map.loc[ent["entity_id"]]
            aum  = float(s.get("aum", s.get("invested_capital", 0)))
            pnl  = float(s.get("pnl_total", s.get("open_pnl",   0)))
            rows.append([
                ent["name"],
                round(aum,  0),
                round(pnl,  2),
                f"{s.get('pct_1d',  0)*100:.2f}%",
                f"{s.get('pct_7d',  0)*100:.2f}%",
                f"{s.get('pct_30d', 0)*100:.2f}%",
                f"{s.get('drawdown', 0)*100:.2f}%",
                f"{s.get('win_rate', 0)*100:.1f}%",
                ent.get("trading_style", ""),
                ent.get("status", ""),
            ])
        return rows

    hdrs   = ["Name","AUM ($)","Open PnL ($)","1D (%)","7D (%)","30D (%)",
              "Max DD (%)","Win Rate (%)","Style","Status"]
    widths = [24, 14, 14, 10, 10, 10, 12, 12, 22, 10]

    for etype, sheet_name in [("pod","Pods"),("strategy","Strategies"),
                               ("trader","Traders"),("venue","Venues")]:
        add_sheet(
            name        = sheet_name,
            col_headers = hdrs,
            data_rows   = _type_rows(etype),
            col_widths  = widths,
            title       = f"Chase Multi-Strategy Portfolio — {sheet_name}",
            subtitle    = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        )

    # --- Equity Curve sheet ---
    curve = ds.get_equity_curve(data_dir)
    port_curve = (curve[curve["entity_id"] == entity_id]
                  .sort_values("timestamp")[["timestamp","equity"]])
    ws_curve = wb.create_sheet("Equity Curve")
    ws_curve.sheet_view.showGridLines = False
    ws_curve.column_dimensions["A"].width = 22
    ws_curve.column_dimensions["B"].width = 16
    for ri, (_, row) in enumerate(port_curve.iterrows(), 1):
        ws_curve.cell(ri, 1, str(row["timestamp"])).font = Font(name="Calibri", size=10, color=EXCEL_TEXT)
        ws_curve.cell(ri, 2, round(row["equity"], 2)).font = Font(name="Calibri", size=10, color=EXCEL_TEXT)
        for ci in (1, 2):
            ws_curve.cell(ri, ci).fill = row_fill(ri)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_summary_sheet(ws, kpis, title_font, meta_font, body_font,
                          hdr_fill, hdr_font, thin_border, left, right, centre,
                          PatternFill, Font, Alignment, BG_DARK, BG_ROW,
                          ACCENT, TEXT, TEXT_MUTED):
    """Write the Portfolio Summary sheet with headline KPIs."""
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 20

    ws.merge_cells("A1:B1")
    ws["A1"].value     = "Chase Multi-Strategy Portfolio — Summary"
    ws["A1"].font      = title_font()
    ws["A1"].fill      = PatternFill("solid", fgColor=BG_DARK)
    ws["A1"].alignment = left()

    ws.merge_cells("A2:B2")
    ws["A2"].value = f"Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws["A2"].font  = meta_font()
    ws["A2"].fill  = PatternFill("solid", fgColor=BG_DARK)

    metrics = [
        ("Initial Investment",  f"${kpis['initial_investment']:,.0f}"),
        ("Current Equity",      f"${kpis['current_equity']:,.0f}"),
        ("Performance (SI)",    f"{kpis['performance']*100:+.2f}%"),
        ("Total Open PnL",      f"${kpis['total_pnl']:,.0f}"),
        ("Return 24h",          f"{kpis['pct_1d']*100:+.2f}%"),
        ("Return 7d",           f"{kpis['pct_7d']*100:+.2f}%"),
        ("Return 30d",          f"{kpis['pct_30d']*100:+.2f}%"),
    ]
    for ri, (label, value) in enumerate(metrics, 4):
        fill = PatternFill("solid", fgColor=BG_ROW)
        ws.cell(ri, 1, label).fill = fill
        ws.cell(ri, 2, value).fill = fill
        ws.cell(ri, 1).font = Font(name="Calibri", size=10, color=TEXT_MUTED)
        ws.cell(ri, 2).font = Font(name="Calibri", size=10, bold=True, color=TEXT)
        for ci in (1, 2):
            ws.cell(ri, ci).border = thin_border()
            ws.cell(ri, ci).alignment = left()


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------

def generate_csv(data_dir: str) -> bytes:
    """Return the full latest snapshots merged with entity metadata as CSV bytes."""
    snaps    = ds.get_snapshots(data_dir)
    entities = ds.get_entities(data_dir)
    latest   = snaps.sort_values("timestamp").groupby("entity_id").last().reset_index()
    merged   = latest.merge(entities[["entity_id","name","entity_type",
                                       "trading_style","status"]], on="entity_id", how="left")
    return merged.to_csv(index=False).encode()


# ---------------------------------------------------------------------------
# PDF export  (basic institutional summary via ReportLab)
# ---------------------------------------------------------------------------

def generate_pdf(data_dir: str) -> bytes:
    """Generate a clean investor-facing PDF summary using ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                    Table, TableStyle, HRFlowable)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             leftMargin=2*cm, rightMargin=2*cm,
                             topMargin=2.5*cm, bottomMargin=2*cm)

    # Colours
    C_BG     = colors.HexColor("#0B1420")
    C_ACCENT = colors.HexColor("#0EA5E9")
    C_TEXT   = colors.HexColor("#E2E8F0")
    C_MUTED  = colors.HexColor("#64748B")
    C_POS    = colors.HexColor("#34D399")
    C_NEG    = colors.HexColor("#F87171")
    C_HEADER = colors.HexColor("#0D1B2E")
    C_ROW    = colors.HexColor("#111C2B")
    C_ALT    = colors.HexColor("#162032")

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Normal"],
                                  fontSize=18, textColor=C_ACCENT,
                                  fontName="Helvetica-Bold", spaceAfter=4)
    sub_style   = ParagraphStyle("sub",   parent=styles["Normal"],
                                  fontSize=9,  textColor=C_MUTED,
                                  fontName="Helvetica",     spaceAfter=12)
    section_style = ParagraphStyle("sec", parent=styles["Normal"],
                                    fontSize=11, textColor=C_TEXT,
                                    fontName="Helvetica-Bold", spaceBefore=14, spaceAfter=6)
    body_style  = ParagraphStyle("body", parent=styles["Normal"],
                                  fontSize=9,  textColor=C_TEXT,
                                  fontName="Helvetica")

    story = []

    # Title
    story.append(Paragraph("Chase Multi-Strategy Portfolio", title_style))
    story.append(Paragraph(
        f"Investor Report  ·  Generated {datetime.utcnow().strftime('%d %B %Y, %H:%M UTC')}",
        sub_style))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_ACCENT))
    story.append(Spacer(1, 10))

    # Portfolio KPIs
    snap_map = (ds.get_snapshots(data_dir)
                  .sort_values("timestamp")
                  .groupby("entity_id").last())

    if "portfolio_main" in snap_map.index:
        kpis = ds.snapshot_kpis(snap_map.loc["portfolio_main"])
        story.append(Paragraph("Portfolio Summary", section_style))

        kpi_data = [
            ["Metric", "Value"],
            ["Initial Investment",  f"${kpis['initial_investment']:>12,.0f}"],
            ["Current Equity",      f"${kpis['current_equity']:>12,.0f}"],
            ["Performance (SI)",    f"{kpis['performance']*100:>+.2f}%"],
            ["Total Open PnL",      f"${kpis['total_pnl']:>12,.0f}"],
            ["Return 24h",          f"{kpis['pct_1d']*100:>+.2f}%"],
            ["Return 7d",           f"{kpis['pct_7d']*100:>+.2f}%"],
            ["Return 30d",          f"{kpis['pct_30d']*100:>+.2f}%"],
        ]
        ts = Table(kpi_data, colWidths=[9*cm, 7*cm])
        ts.setStyle(TableStyle([
            ("BACKGROUND",   (0,0), (-1,0),  C_HEADER),
            ("TEXTCOLOR",    (0,0), (-1,0),  C_ACCENT),
            ("FONTNAME",     (0,0), (-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",     (0,0), (-1,-1), 9),
            ("BACKGROUND",   (0,1), (-1,-1), C_ROW),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_ROW, C_ALT]),
            ("TEXTCOLOR",    (0,1), (-1,-1), C_TEXT),
            ("ALIGN",        (1,0), (1,-1),  "RIGHT"),
            ("LEFTPADDING",  (0,0), (-1,-1), 8),
            ("RIGHTPADDING", (0,0), (-1,-1), 8),
            ("TOPPADDING",   (0,0), (-1,-1), 5),
            ("BOTTOMPADDING",(0,0), (-1,-1), 5),
            ("LINEBELOW",    (0,0), (-1,0),  0.5, C_ACCENT),
        ]))
        story.append(ts)
        story.append(Spacer(1, 14))

    # Pod breakdown
    entities = ds.get_entities(data_dir)
    pods     = entities[entities["entity_type"] == "pod"]
    if not pods.empty:
        story.append(Paragraph("Pod Overview", section_style))
        pod_data = [["Pod", "Invested ($)", "Equity ($)", "PnL ($)", "7d (%)", "30d (%)"]]
        for _, pod in pods.iterrows():
            if pod["entity_id"] not in snap_map.index:
                continue
            s = snap_map.loc[pod["entity_id"]]
            aum = float(s.get("aum", 0))
            pod_data.append([
                pod["name"],
                f"{aum:,.0f}",
                f"{float(s['equity']):,.0f}",
                f"{float(s.get('pnl_total', s.get('open_pnl', 0))):,.0f}",
                f"{float(s.get('pct_7d', 0))*100:+.2f}%",
                f"{float(s.get('pct_30d',0))*100:+.2f}%",
            ])
        pt = Table(pod_data, colWidths=[6*cm,3*cm,3*cm,2.5*cm,2.5*cm,2.5*cm])
        pt.setStyle(TableStyle([
            ("BACKGROUND",    (0,0),(-1,0),  C_HEADER),
            ("TEXTCOLOR",     (0,0),(-1,0),  C_ACCENT),
            ("FONTNAME",      (0,0),(-1,0),  "Helvetica-Bold"),
            ("FONTSIZE",      (0,0),(-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1), [C_ROW, C_ALT]),
            ("TEXTCOLOR",     (0,1),(-1,-1), C_TEXT),
            ("ALIGN",         (1,0),(-1,-1), "RIGHT"),
            ("LEFTPADDING",   (0,0),(-1,-1), 6),
            ("RIGHTPADDING",  (0,0),(-1,-1), 6),
            ("TOPPADDING",    (0,0),(-1,-1), 4),
            ("BOTTOMPADDING", (0,0),(-1,-1), 4),
        ]))
        story.append(pt)

    story.append(Spacer(1, 20))
    story.append(HRFlowable(width="100%", thickness=0.5, color=C_MUTED))
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "This report is for authorised personnel only. "
        "Past performance is not indicative of future results.",
        ParagraphStyle("disc", parent=styles["Normal"], fontSize=7,
                        textColor=C_MUTED, fontName="Helvetica")))

    doc.build(story)
    return buf.getvalue()
