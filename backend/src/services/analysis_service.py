# backend/src/services/analysis_service.py
"""
AXIA Trade Analysis Service
Parses AXIA statement Excel files and computes full trade analytics.
Generates professional multi-sheet Excel export reports.
"""

import io
import os
import uuid
from collections import defaultdict
from datetime import datetime, timedelta
from typing import Optional

import requests
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# OANDA REST API credentials (GBP rate conversion)
_OANDA_URL   = os.getenv('OANDA_API_URL', 'https://api-fxpractice.oanda.com/v3')
_OANDA_TOKEN = os.getenv('OANDA_TOKEN',   '37ee33b35f88e073a08d533849f7a24b-524c89ef15f36cfe532f0918a6aee4c2')

# (OANDA pair, invert) for CCY -> GBP
# invert=True  : rate_to_gbp = 1/close  (GBP_USD: 1 GBP = X USD -> USD->GBP = 1/X)
# invert=False : rate_to_gbp = close    (EUR_GBP: 1 EUR = X GBP -> EUR->GBP = X)
_CCY_GBP_PAIR: dict[str, tuple] = {
    'USD': ('GBP_USD', True),
    'EUR': ('EUR_GBP', False),
    'JPY': ('GBP_JPY', True),
    'CAD': ('GBP_CAD', True),
    'CHF': ('GBP_CHF', True),
    'AUD': ('GBP_AUD', True),
    'NZD': ('GBP_NZD', True),
    'HKD': ('GBP_HKD', True),
    'CNH': ('GBP_CNH', True),
    'SGD': ('GBP_SGD', True),
}

# --- In-memory analysis cache (analysis_id -> {data, created_at}) ------------
_cache: dict[str, dict] = {}


# --- Parse -------------------------------------------------------------------

def parse_statement(file_bytes: bytes) -> dict:
    """
    Parse AXIA statement Excel file and return full analysis dict.
    Expected columns: TRADE DATE, DELIVERY / PRODUCT, LONG, SHORT,
    REALIZED PnL, COMMISION FEES, MARKET FEES, NFA FEES, TOTAL COMMS,
    TOTAL PnL, CURRENCY
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    all_rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]
    if not all_rows:
        raise ValueError("Empty spreadsheet")

    header = [str(c).strip() if c else '' for c in all_rows[0]]
    raw: list[dict] = []

    for row in all_rows[1:]:
        r = dict(zip(header, row))
        td = r.get('TRADE DATE')
        if td is None:
            continue
        if hasattr(td, 'date'):
            td = td.date()
        elif isinstance(td, str):
            td = datetime.strptime(td[:10], '%Y-%m-%d').date()

        raw.append({
            'date':            str(td),
            'instrument':      str(r.get('DELIVERY / PRODUCT', '')).strip(),
            'currency':        str(r.get('CURRENCY', 'USD')).strip(),
            'long':            float(r.get('LONG')            or 0),
            'short':           float(r.get('SHORT')           or 0),
            'realized_pnl':    float(r.get('REALIZED PnL')    or 0),
            'commission_fees': float(r.get('COMMISION FEES')  or 0),
            'market_fees':     float(r.get('MARKET FEES')     or 0),
            'nfa_fees':        float(r.get('NFA FEES')        or 0),
            'total_comms':     float(r.get('TOTAL COMMS')     or 0),
            'net_pnl':         float(r.get('TOTAL PnL')       or 0),
        })

    if not raw:
        raise ValueError("No valid data rows found in spreadsheet")

    # -- Aggregate by instrument -----------------------------------------------
    asset_map: dict[str, dict] = {}
    for r in raw:
        key = f"{r['instrument']}|{r['currency']}"
        if key not in asset_map:
            asset_map[key] = {
                'instrument':      r['instrument'],
                'currency':        r['currency'],
                'long':            0.0,
                'short':           0.0,
                'realized_pnl':    0.0,
                'commission_fees': 0.0,
                'market_fees':     0.0,
                'nfa_fees':        0.0,
                'total_comms':     0.0,
                'net_pnl':         0.0,
                'trade_days':      0,
                'daily':           [],
            }
        a = asset_map[key]
        a['long']            += r['long']
        a['short']           += r['short']
        a['realized_pnl']    += r['realized_pnl']
        a['commission_fees'] += r['commission_fees']
        a['market_fees']     += r['market_fees']
        a['nfa_fees']        += r['nfa_fees']
        a['total_comms']     += r['total_comms']
        a['net_pnl']         += r['net_pnl']
        a['trade_days']      += 1
        a['daily'].append({
            'date':         r['date'],
            'long':         r['long'],
            'short':        r['short'],
            'realized_pnl': r['realized_pnl'],
            'total_comms':  r['total_comms'],
            'net_pnl':      r['net_pnl'],
        })

    # Compute per-asset cumulative equity and round floats
    assets = list(asset_map.values())
    for a in assets:
        a['daily'] = sorted(a['daily'], key=lambda x: x['date'])
        cum = 0.0
        for d in a['daily']:
            cum += d['net_pnl']
            d['cumulative'] = round(cum, 2)
        # Round all float fields
        for field in ('long', 'short', 'realized_pnl', 'commission_fees',
                      'market_fees', 'nfa_fees', 'total_comms', 'net_pnl'):
            a[field] = round(a[field], 2)

    # Build portfolio_daily_detail: sorted by date ASC then instrument
    # Cumulative tracks portfolio running total per currency across all instruments
    ccy_running: dict[str, float] = defaultdict(float)
    portfolio_daily_detail = []
    for r in sorted(raw, key=lambda x: (x['date'], x['instrument'])):
        ccy_running[r['currency']] += r['net_pnl']
        portfolio_daily_detail.append({
            'date':         r['date'],
            'instrument':   r['instrument'],
            'currency':     r['currency'],
            'long':         r['long'],
            'short':        r['short'],
            'realized_pnl': r['realized_pnl'],
            'total_comms':  r['total_comms'],
            'net_pnl':      r['net_pnl'],
            'cumulative':   round(ccy_running[r['currency']], 2),
        })

    # Sort by net_pnl ascending (worst first -- canonical order)
    assets.sort(key=lambda x: x['net_pnl'])

    # -- Aggregate by date -----------------------------------------------------
    date_map: dict[str, dict] = defaultdict(lambda: defaultdict(float))
    for r in raw:
        ccy = r['currency']
        date_map[r['date']][f'{ccy}_pnl']   += r['net_pnl']
        date_map[r['date']][f'{ccy}_comms'] += r['total_comms']
        date_map[r['date']]['total_lots']    += r['long'] + r['short']
        # Track per-date instrument breakdown for heatmap
        key = f'inst_{r["instrument"]}|{ccy}'
        date_map[r['date']][key] = round(date_map[r['date']][key] + r['net_pnl'], 2)

    all_dates = sorted(date_map.keys())

    # Per-currency cumulative equity across dates
    currencies = sorted({r['currency'] for r in raw})
    ccy_cum = {c: 0.0 for c in currencies}
    by_date = []
    for d in all_dates:
        row = {'date': d}
        row.update(date_map[d])
        for c in currencies:
            ccy_cum[c] += row.get(f'{c}_pnl', 0.0)
            row[f'{c}_cum'] = round(ccy_cum[c], 2)
        # Round daily totals
        for k, v in row.items():
            if isinstance(v, float):
                row[k] = round(v, 2)
        by_date.append(row)

    # -- Currency summary ------------------------------------------------------
    by_currency: dict[str, dict] = defaultdict(
        lambda: {'net_pnl': 0.0, 'total_comms': 0.0, 'realized_pnl': 0.0}
    )
    for a in assets:
        c = a['currency']
        by_currency[c]['net_pnl']      += a['net_pnl']
        by_currency[c]['total_comms']  += a['total_comms']
        by_currency[c]['realized_pnl'] += a['realized_pnl']
    by_currency_clean = {
        k: {kk: round(vv, 2) for kk, vv in v.items()}
        for k, v in by_currency.items()
    }

    # -- Commission totals -----------------------------------------------------
    total_broker = sum(a['commission_fees'] for a in assets)
    total_mkt    = sum(a['market_fees']     for a in assets)
    total_nfa    = sum(a['nfa_fees']        for a in assets)

    # -- Commission drag per asset ---------------------------------------------
    # (% of gross realized PnL lost to commissions -- useful for manager report)
    for a in assets:
        gross = abs(a['realized_pnl'])
        comms = abs(a['total_comms'])
        a['comm_drag_pct'] = round((comms / gross * 100) if gross > 0 else 0, 1)

    # -- GBP view via OANDA daily close rates ----------------------------------
    _gbp_lookup, _gbp_meta = _build_rate_lookup(raw)
    _gbp_view               = _compute_gbp_view(raw, _gbp_lookup, _gbp_meta)

    return {
        'date_range': {
            'from':         all_dates[0],
            'to':           all_dates[-1],
            'trading_days': len(all_dates),
            'dates':        all_dates,
        },
        'summary': {
            'total_instruments': len(assets),
            'currencies':        currencies,
            'total_long_lots':   round(sum(a['long']  for a in assets), 0),
            'total_short_lots':  round(sum(a['short'] for a in assets), 0),
            'total_trades':      len(raw),
        },
        'by_currency':  by_currency_clean,
        'by_asset':     assets,
        'by_date':      by_date,
        'commission_breakdown': {
            'commission_fees': round(total_broker, 2),
            'market_fees':     round(total_mkt, 2),
            'nfa_fees':        round(total_nfa, 2),
            'total':           round(total_broker + total_mkt + total_nfa, 2),
        },
        'top_winners':            sorted(assets, key=lambda x: x['net_pnl'], reverse=True)[:5],
        'top_losers':             sorted(assets, key=lambda x: x['net_pnl'])[:5],
        'portfolio_daily_detail': portfolio_daily_detail,
        '_raw':                   raw,   # kept for /refresh-gbp; stripped before API response
        **_gbp_view,
    }


# --- GBP Rate Fetching -------------------------------------------------------

def _fetch_oanda_pair_series(pair: str, from_date: str, to_date: str) -> dict[str, float]:
    """
    Fetch daily close mid-prices from OANDA for `pair` over the date range.
    Returns {date_str: close_mid}. Returns {} on any error (graceful degradation).
    """
    try:
        from_dt = (datetime.strptime(from_date, '%Y-%m-%d') - timedelta(days=4)).strftime('%Y-%m-%dT00:00:00Z')
        to_dt   = (datetime.strptime(to_date,   '%Y-%m-%d') + timedelta(days=2)).strftime('%Y-%m-%dT23:59:59Z')
        url = f"{_OANDA_URL}/instruments/{pair}/candles"
        r   = requests.get(
            url,
            headers={'Authorization': f'Bearer {_OANDA_TOKEN}'},
            params={'granularity': 'D', 'price': 'M', 'from': from_dt, 'to': to_dt},
            timeout=15,
        )
        if r.status_code != 200:
            return {}
        result = {}
        for c in r.json().get('candles', []):
            date = c['time'][:10]
            result[date] = float(c['mid']['c'])
        return result
    except Exception:
        return {}


def _build_rate_lookup(raw: list[dict]) -> tuple[dict[tuple, float], dict]:
    """
    Build a {(date, ccy): gbp_rate} lookup from OANDA daily closes.
    Backward-fills weekends / holidays with the last available rate.

    Returns (lookup, meta) where meta = {ok, fetched, failed, unknown}.
    - ok      : True only if every non-GBP currency got real rates
    - fetched : currencies with successful OANDA fetch
    - failed  : currencies whose OANDA pair returned no data
    - unknown : currencies with no known OANDA pair at all
    """
    if not raw:
        return {}, {'ok': True, 'fetched': [], 'failed': [], 'unknown': []}

    dates      = sorted({r['date'] for r in raw})
    currencies = sorted({r['currency'] for r in raw})
    from_date  = dates[0]
    to_date    = dates[-1]

    # Fetch series once per required OANDA pair
    pair_cache: dict[str, dict[str, float]] = {}
    for ccy in currencies:
        if ccy == 'GBP' or ccy not in _CCY_GBP_PAIR:
            continue
        pair, _ = _CCY_GBP_PAIR[ccy]
        if pair not in pair_cache:
            pair_cache[pair] = _fetch_oanda_pair_series(pair, from_date, to_date)

    # Track per-currency fetch status
    fetched: list[str] = []
    failed:  list[str] = []
    unknown: list[str] = []

    lookup: dict[tuple, float] = {}
    for ccy in currencies:
        if ccy == 'GBP':
            for d in dates:
                lookup[(d, ccy)] = 1.0
            continue

        if ccy not in _CCY_GBP_PAIR:
            unknown.append(ccy)
            # Store sentinel so caller knows rate is missing
            for d in dates:
                lookup[(d, ccy)] = None   # type: ignore[assignment]
            continue

        pair, invert = _CCY_GBP_PAIR[ccy]
        series       = pair_cache.get(pair, {})
        available    = sorted(series.keys())

        if not series:
            failed.append(ccy)
            for d in dates:
                lookup[(d, ccy)] = None   # type: ignore[assignment]
            continue

        fetched.append(ccy)
        for d in dates:
            raw_rate = None
            for ad in reversed(available):
                if ad <= d:
                    raw_rate = series[ad]
                    break
            if raw_rate is None and available:
                raw_rate = series[available[0]]
            gbp_rate = (1.0 / raw_rate if invert else raw_rate) if raw_rate else None
            lookup[(d, ccy)] = round(gbp_rate, 8) if gbp_rate else None  # type: ignore[assignment]

    all_ok = len(failed) == 0 and len(unknown) == 0
    return lookup, {'ok': all_ok, 'fetched': fetched, 'failed': failed + unknown, 'unknown': unknown}


def _compute_gbp_view(raw: list[dict], lookup: dict[tuple, float], rates_meta: dict) -> dict:
    """
    Convert all raw trade rows to GBP using the rate lookup, then rebuild
    the full aggregation set (same structure as parse_statement output).
    Returns a dict of gbp_* keys to merge into the main analysis result.
    """
    # Per-row GBP conversion -- only include rows where a real rate exists
    gbp_raw = []
    for r in raw:
        rate = lookup.get((r['date'], r['currency']))
        if rate is None:
            continue   # skip rows with missing rates -- no false data
        gbp_raw.append({
            'date':              r['date'],
            'instrument':        r['instrument'],
            'currency':          'GBP',
            'original_currency': r['currency'],
            'gbp_rate':          rate,
            'long':              r['long'],
            'short':             r['short'],
            # Native values preserved for FX Rate Log sheet
            'realized_pnl_native': round(r['realized_pnl'], 2),
            'total_comms_native':  round(r['total_comms'],  2),
            'net_pnl_native':      round(r['net_pnl'],      2),
            # GBP-converted values
            'realized_pnl':      round(r['realized_pnl']    * rate, 2),
            'commission_fees':   round(r['commission_fees']  * rate, 2),
            'market_fees':       round(r['market_fees']      * rate, 2),
            'nfa_fees':          round(r['nfa_fees']         * rate, 2),
            'total_comms':       round(r['total_comms']      * rate, 2),
            'net_pnl':           round(r['net_pnl']          * rate, 2),
        })

    # Aggregate by instrument
    asset_map: dict[str, dict] = {}
    for r in gbp_raw:
        key = r['instrument']
        if key not in asset_map:
            asset_map[key] = {
                'instrument':        r['instrument'],
                'currency':          'GBP',
                'original_currency': r['original_currency'],
                'long':              0.0, 'short':           0.0,
                'realized_pnl':      0.0, 'commission_fees': 0.0,
                'market_fees':       0.0, 'nfa_fees':        0.0,
                'total_comms':       0.0, 'net_pnl':         0.0,
                'trade_days':        0,   'daily':           [],
            }
        a = asset_map[key]
        a['long']            += r['long']
        a['short']           += r['short']
        a['realized_pnl']    += r['realized_pnl']
        a['commission_fees'] += r['commission_fees']
        a['market_fees']     += r['market_fees']
        a['nfa_fees']        += r['nfa_fees']
        a['total_comms']     += r['total_comms']
        a['net_pnl']         += r['net_pnl']
        a['trade_days']      += 1
        a['daily'].append({
            'date':         r['date'],
            'long':         r['long'],
            'short':        r['short'],
            'realized_pnl': r['realized_pnl'],
            'total_comms':  r['total_comms'],
            'net_pnl':      r['net_pnl'],
        })

    assets = list(asset_map.values())
    for a in assets:
        a['daily'] = sorted(a['daily'], key=lambda x: x['date'])
        cum = 0.0
        for d in a['daily']:
            cum += d['net_pnl']
            d['cumulative'] = round(cum, 2)
        for field in ('long', 'short', 'realized_pnl', 'commission_fees',
                      'market_fees', 'nfa_fees', 'total_comms', 'net_pnl'):
            a[field] = round(a[field], 2)
        gross = abs(a['realized_pnl'])
        comms = abs(a['total_comms'])
        a['comm_drag_pct'] = round((comms / gross * 100) if gross > 0 else 0, 1)

    assets.sort(key=lambda x: x['net_pnl'])

    # Portfolio daily detail in GBP (single running cumulative)
    gbp_running = 0.0
    gbp_portfolio_daily = []
    for r in sorted(gbp_raw, key=lambda x: (x['date'], x['instrument'])):
        gbp_running += r['net_pnl']
        gbp_portfolio_daily.append({
            'date':              r['date'],
            'instrument':        r['instrument'],
            'currency':          'GBP',
            'original_currency': r['original_currency'],
            'gbp_rate':          r['gbp_rate'],
            'long':              r['long'],
            'short':             r['short'],
            # Native original values (for FX Rate Log Excel sheet)
            'realized_pnl_native': r['realized_pnl_native'],
            'total_comms_native':  r['total_comms_native'],
            'net_pnl_native':      r['net_pnl_native'],
            # GBP values
            'realized_pnl':      r['realized_pnl'],
            'total_comms':       r['total_comms'],
            'net_pnl':           r['net_pnl'],
            'cumulative':        round(gbp_running, 2),
        })

    # Daily totals in GBP
    date_map: dict[str, dict] = defaultdict(lambda: defaultdict(float))
    for r in gbp_raw:
        date_map[r['date']]['GBP_pnl']   += r['net_pnl']
        date_map[r['date']]['GBP_comms'] += r['total_comms']
        date_map[r['date']]['total_lots'] += r['long'] + r['short']

    all_dates  = sorted(date_map.keys())
    gbp_cum    = 0.0
    gbp_by_date = []
    for d in all_dates:
        row = {'date': d}
        row.update(date_map[d])
        gbp_cum += row.get('GBP_pnl', 0.0)
        row['GBP_cum'] = round(gbp_cum, 2)
        for k, v in row.items():
            if isinstance(v, float):
                row[k] = round(v, 2)
        gbp_by_date.append(row)

    # Currency summary (single GBP bucket)
    total_realized = round(sum(r['realized_pnl'] for r in gbp_raw), 2)
    total_comms    = round(sum(r['total_comms']  for r in gbp_raw), 2)
    total_net      = round(sum(r['net_pnl']      for r in gbp_raw), 2)
    gbp_by_currency = {
        'GBP': {
            'net_pnl':      total_net,
            'total_comms':  total_comms,
            'realized_pnl': total_realized,
        }
    }

    gbp_commission_breakdown = {
        'commission_fees': round(sum(r['commission_fees'] for r in gbp_raw), 2),
        'market_fees':     round(sum(r['market_fees']     for r in gbp_raw), 2),
        'nfa_fees':        round(sum(r['nfa_fees']        for r in gbp_raw), 2),
        'total':           total_comms,
    }

    # Base status -- always returned so frontend knows whether toggle is safe
    result: dict = {
        'gbp_rates_ok':      rates_meta['ok'],
        'gbp_rates_fetched': rates_meta['fetched'],
        'gbp_rates_failed':  rates_meta['failed'],
    }

    # Only attach GBP data when every required currency had a real rate.
    # This prevents any false results from 1.0 fallbacks reaching the UI.
    if rates_meta['ok'] and gbp_raw:
        result.update({
            'gbp_assets':               assets,
            'gbp_by_date':              gbp_by_date,
            'gbp_by_currency':          gbp_by_currency,
            'gbp_portfolio_daily':      gbp_portfolio_daily,
            'gbp_commission_breakdown': gbp_commission_breakdown,
            'gbp_top_winners':          sorted(assets, key=lambda x: x['net_pnl'], reverse=True)[:5],
            'gbp_top_losers':           sorted(assets, key=lambda x: x['net_pnl'])[:5],
            'gbp_summary': {
                'total_net_gbp':   total_net,
                'total_comms_gbp': total_comms,
            },
        })

    return result


# --- Cache --------------------------------------------------------------------

def store_analysis(data: dict) -> str:
    aid = str(uuid.uuid4())
    _cache[aid] = {'data': data, 'created_at': datetime.utcnow()}
    if len(_cache) > 50:
        oldest = sorted(_cache, key=lambda k: _cache[k]['created_at'])[:10]
        for k in oldest:
            del _cache[k]
    return aid


def get_analysis(analysis_id: str) -> Optional[dict]:
    entry = _cache.get(analysis_id)
    return entry['data'] if entry else None


def refresh_gbp_rates(analysis_id: str) -> dict:
    """
    Re-fetch OANDA GBP rates for a cached analysis and update the cache.
    Returns the updated gbp_* fields.
    Raises ValueError if analysis not found or raw data unavailable.
    Raises RuntimeError if OANDA fetch still fails.
    """
    data = get_analysis(analysis_id)
    if not data:
        raise ValueError("Analysis not found or expired -- please re-upload the file")

    raw = data.get('_raw')
    if not raw:
        raise ValueError("Raw data unavailable -- please re-upload the file")

    lookup, meta = _build_rate_lookup(raw)

    if not meta['ok']:
        failed_str = ', '.join(meta['failed']) if meta['failed'] else 'unknown pairs'
        raise RuntimeError(
            f"OANDA rate fetch failed for: {failed_str}. "
            "Check API connectivity and try again."
        )

    gbp_view = _compute_gbp_view(raw, lookup, meta)

    # Merge updated GBP fields into the cached data
    data.update(gbp_view)

    return gbp_view


# --- Excel Export Helpers -----------------------------------------------------

_NAVY       = '1F3864'
_MID_NAVY   = '2E4F7A'
_LIGHT_BLUE = 'D6E4F0'
_POS_FILL   = 'C6EFCE'
_NEG_FILL   = 'FFC7CE'
_ALT_FILL   = 'EBF2F8'
_WHITE      = 'FFFFFF'
_POS_FONT   = '006100'
_NEG_FONT   = '9C0006'


def _hdr_font(size: int = 10) -> Font:
    return Font(bold=True, color='FFFFFF', name='Calibri', size=size)


def _hdr_fill() -> PatternFill:
    return PatternFill('solid', fgColor=_NAVY)


def _sub_hdr_fill() -> PatternFill:
    return PatternFill('solid', fgColor=_MID_NAVY)


def _pos_fill() -> PatternFill:
    return PatternFill('solid', fgColor=_POS_FILL)


def _neg_fill() -> PatternFill:
    return PatternFill('solid', fgColor=_NEG_FILL)


def _alt_fill() -> PatternFill:
    return PatternFill('solid', fgColor=_ALT_FILL)


def _white_fill() -> PatternFill:
    return PatternFill('solid', fgColor=_WHITE)


def _ctr() -> Alignment:
    return Alignment(horizontal='center', vertical='center', wrap_text=False)


def _right() -> Alignment:
    return Alignment(horizontal='right', vertical='center')


def _left() -> Alignment:
    return Alignment(horizontal='left', vertical='center')


def _thin_border() -> Border:
    s = Side(style='thin', color='B8CCE4')
    return Border(left=s, right=s, top=s, bottom=s)


def _write_headers(ws, row: int, cols: list[str], start_col: int = 1,
                   widths: list[float] = None, fill=None):
    """Write a styled header row."""
    for i, h in enumerate(cols):
        col = start_col + i
        c = ws.cell(row=row, column=col, value=h)
        c.font      = _hdr_font()
        c.fill      = fill() if fill else _hdr_fill()
        c.alignment = _ctr()
        c.border    = _thin_border()
    if widths:
        for i, w in enumerate(widths):
            ws.column_dimensions[get_column_letter(start_col + i)].width = w


def _style_pnl_cell(cell, value: float):
    """Apply green/red fill + font to a P&L cell."""
    if value >= 0:
        cell.fill = _pos_fill()
        cell.font = Font(bold=True, color=_POS_FONT, name='Calibri', size=10)
    else:
        cell.fill = _neg_fill()
        cell.font = Font(bold=True, color=_NEG_FONT, name='Calibri', size=10)


def _fmt_num(ws, row: int, col: int, value: float, pnl: bool = False,
             bold: bool = False, alt: bool = False):
    c = ws.cell(row=row, column=col, value=round(value, 2))
    c.alignment = _right()
    c.border    = _thin_border()
    c.number_format = '#,##0.00'
    if pnl:
        _style_pnl_cell(c, value)
    else:
        c.fill = _alt_fill() if alt else _white_fill()
        if bold:
            c.font = Font(bold=True, name='Calibri', size=10)
    return c


def _txt(ws, row: int, col: int, value, bold: bool = False, color: str = None,
         align: str = 'left', alt: bool = False):
    c = ws.cell(row=row, column=col, value=value)
    c.border    = _thin_border()
    c.alignment = _left() if align == 'left' else _ctr()
    c.fill      = _alt_fill() if alt else _white_fill()
    f = Font(name='Calibri', size=10)
    if bold:
        f = Font(name='Calibri', size=10, bold=True)
    if color:
        f = Font(name='Calibri', size=10, bold=bold, color=color)
    c.font = f
    return c


# --- GBP Excel Export --------------------------------------------------------

def _generate_gbp_excel_report(data: dict, trader: str, account: str) -> bytes:
    """
    Generate a GBP-view Excel report.
    All P&L / commission figures are converted to GBP via OANDA daily close rates.
    Includes an FX Rate Log sheet showing native values, rate used, and GBP result.
    """
    wb = openpyxl.Workbook()

    fetched = data.get("gbp_rates_fetched", [])
    gbp_assets = data.get("gbp_assets", [])
    gbp_by_date = data.get("gbp_by_date", [])
    gbp_by_currency = data.get("gbp_by_currency", {"GBP": {"net_pnl": 0, "total_comms": 0, "realized_pnl": 0}})
    gbp_cb = data.get("gbp_commission_breakdown", {"commission_fees": 0, "market_fees": 0, "nfa_fees": 0, "total": 0})
    gbp_daily = data.get("gbp_portfolio_daily", [])
    gbp_winners = data.get("gbp_top_winners", [])
    gbp_losers  = data.get("gbp_top_losers",  [])

    # Sheet 1: Executive Summary (GBP) ----------------------------------------
    ws1 = wb.active
    ws1.title = "Executive Summary (GBP)"
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = "1F3864"

    ws1.merge_cells("A1:K1")
    t = ws1["A1"]
    t.value     = "AXIA STRATEGY -- TRADE ANALYSIS REPORT  |  GBP VIEW"
    t.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=16)
    t.fill      = _hdr_fill()
    t.alignment = _ctr()
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:K2")
    s = ws1["A2"]
    s.value = (
        f"Trader: {trader}   |   Account: {account}   |   "
        f"Period: {data['date_range']['from']} -> {data['date_range']['to']}   |   "
        f"Currencies converted to GBP via OANDA daily close rates   |   "
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    )
    s.font      = Font(italic=True, color="FFFFFF", name="Calibri", size=10)
    s.fill      = PatternFill("solid", fgColor=_MID_NAVY)
    s.alignment = _ctr()
    ws1.row_dimensions[2].height = 20

    # Period summary
    ws1.cell(row=4, column=1, value="PERIOD SUMMARY").font = Font(bold=True, name="Calibri", size=11, color=_NAVY)
    period_rows = [
        ("Date From",          data["date_range"]["from"]),
        ("Date To",            data["date_range"]["to"]),
        ("Trading Days",       data["date_range"]["trading_days"]),
        ("Instruments Traded", data["summary"]["total_instruments"]),
        ("Total Long Lots",    int(data["summary"]["total_long_lots"])),
        ("Total Short Lots",   int(data["summary"]["total_short_lots"])),
        ("Trade Rows",         data["summary"]["total_trades"]),
        ("Native Currencies",  ", ".join(data["summary"]["currencies"])),
        ("Rates Fetched For",  ", ".join(fetched) if fetched else "N/A"),
    ]
    _write_headers(ws1, 5, ["Metric", "Value"], start_col=1, widths=[24, 20])
    for i, (label, val) in enumerate(period_rows, 6):
        alt = (i % 2 == 0)
        _txt(ws1, i, 1, label, bold=True, alt=alt)
        _txt(ws1, i, 2, val, alt=alt)

    # GBP P&L Summary
    ws1.cell(row=4, column=4, value="GBP P&L SUMMARY").font = Font(bold=True, name="Calibri", size=11, color=_NAVY)
    _write_headers(ws1, 5, ["", "Realized P&L", "Total Comms", "Net P&L"],
                   start_col=4, widths=[14, 16, 16, 16])
    gbp_totals = gbp_by_currency.get("GBP", {})
    _txt(ws1, 6, 4, "GBP TOTAL", bold=True)
    _fmt_num(ws1, 6, 5, gbp_totals.get("realized_pnl", 0))
    _fmt_num(ws1, 6, 6, gbp_totals.get("total_comms",  0))
    _fmt_num(ws1, 6, 7, gbp_totals.get("net_pnl",      0), pnl=True)

    # Commission Breakdown (GBP)
    ws1.cell(row=4, column=9, value="COMMISSION BREAKDOWN (GBP)").font = Font(bold=True, name="Calibri", size=11, color=_NAVY)
    _write_headers(ws1, 5, ["Fee Type", "Amount (GBP)"], start_col=9, widths=[24, 18])
    comm_rows = [
        ("Broker / Exchange Commission", gbp_cb["commission_fees"]),
        ("Market / Clearing Fees",       gbp_cb["market_fees"]),
        ("NFA Regulatory Fees",          gbp_cb["nfa_fees"]),
        ("TOTAL COMMISSIONS (GBP)",      gbp_cb["total"]),
    ]
    for i, (label, val) in enumerate(comm_rows, 6):
        alt = (i % 2 == 0)
        bold = label.startswith("TOTAL")
        _txt(ws1, i, 9, label, bold=bold, alt=alt)
        _fmt_num(ws1, i, 10, val, bold=bold, alt=alt)

    # Top winners / losers (GBP)
    ws1.cell(row=14, column=1, value="* TOP PERFORMERS (GBP)").font = Font(bold=True, name="Calibri", size=11, color="006100")
    _write_headers(ws1, 15, ["Instrument", "Orig CCY", "Long Lots", "Short Lots", "Net P&L (GBP)"], start_col=1)
    for i, a in enumerate(gbp_winners, 16):
        alt = (i % 2 == 0)
        _txt(ws1, i, 1, a["instrument"], alt=alt)
        _txt(ws1, i, 2, a.get("original_currency", ""), alt=alt, align="center")
        _txt(ws1, i, 3, int(a["long"]),  alt=alt, align="center")
        _txt(ws1, i, 4, int(a["short"]), alt=alt, align="center")
        _fmt_num(ws1, i, 5, a["net_pnl"], pnl=True)

    ws1.cell(row=14, column=7, value="v WORST PERFORMERS (GBP)").font = Font(bold=True, name="Calibri", size=11, color="9C0006")
    _write_headers(ws1, 15, ["Instrument", "Orig CCY", "Long Lots", "Short Lots", "Net P&L (GBP)"], start_col=7)
    for i, a in enumerate(gbp_losers, 16):
        alt = (i % 2 == 0)
        _txt(ws1, i, 7,  a["instrument"], alt=alt)
        _txt(ws1, i, 8,  a.get("original_currency", ""), alt=alt, align="center")
        _txt(ws1, i, 9,  int(a["long"]),  alt=alt, align="center")
        _txt(ws1, i, 10, int(a["short"]), alt=alt, align="center")
        _fmt_num(ws1, i, 11, a["net_pnl"], pnl=True)

    ws1.freeze_panes = "A3"

    # Sheet 2: Asset Breakdown (GBP) ------------------------------------------
    ws2 = wb.create_sheet("Asset Breakdown (GBP)")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"
    ws2.sheet_properties.tabColor = _MID_NAVY

    a_cols   = ["Instrument", "Orig CCY", "Long Lots", "Short Lots", "Trade Days",
                "Realized P&L (GBP)", "Broker Comms (GBP)", "Market Fees (GBP)",
                "NFA Fees (GBP)", "Total Comms (GBP)", "Net P&L (GBP)", "Comm Drag %"]
    a_widths = [24, 10, 12, 12, 12, 18, 18, 18, 14, 18, 18, 14]
    _write_headers(ws2, 1, a_cols, widths=a_widths)

    for i, a in enumerate(sorted(gbp_assets, key=lambda x: x["net_pnl"]), 2):
        alt = (i % 2 == 0)
        _txt(ws2, i, 1,  a["instrument"],                  alt=alt)
        _txt(ws2, i, 2,  a.get("original_currency", ""),   alt=alt, align="center")
        _txt(ws2, i, 3,  int(a["long"]),                   alt=alt, align="center")
        _txt(ws2, i, 4,  int(a["short"]),                  alt=alt, align="center")
        _txt(ws2, i, 5,  a["trade_days"],                  alt=alt, align="center")
        _fmt_num(ws2, i, 6,  a["realized_pnl"],    alt=alt)
        _fmt_num(ws2, i, 7,  a["commission_fees"],  alt=alt)
        _fmt_num(ws2, i, 8,  a["market_fees"],      alt=alt)
        _fmt_num(ws2, i, 9,  a["nfa_fees"],         alt=alt)
        _fmt_num(ws2, i, 10, a["total_comms"],      alt=alt)
        _fmt_num(ws2, i, 11, a["net_pnl"],          pnl=True)
        _txt(ws2, i, 12, f"{a['comm_drag_pct']:.1f}%", alt=alt, align="center")

    # GBP total row
    tot_row = len(gbp_assets) + 3
    ws2.cell(row=tot_row, column=1, value="GBP TOTAL").font = Font(bold=True, name="Calibri", size=10)
    _fmt_num(ws2, tot_row, 11, gbp_totals.get("net_pnl", 0),      pnl=True)
    _fmt_num(ws2, tot_row, 10, gbp_totals.get("total_comms", 0),  bold=True)
    _fmt_num(ws2, tot_row, 6,  gbp_totals.get("realized_pnl", 0), bold=True)

    # Sheet 3: Daily P&L (GBP) ------------------------------------------------
    ws3 = wb.create_sheet("Daily P&L (GBP)")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"
    ws3.sheet_properties.tabColor = _MID_NAVY

    _write_headers(ws3, 1, ["Date", "Total Lots", "Net P&L (GBP)", "Total Comms (GBP)", "Cumulative P&L (GBP)"],
                   widths=[14, 12, 18, 20, 22])
    for i, row in enumerate(gbp_by_date, 2):
        alt = (i % 2 == 0)
        _txt(ws3, i, 1, row["date"],                       alt=alt, align="center")
        _txt(ws3, i, 2, int(row.get("total_lots", 0)),     alt=alt, align="center")
        pnl_val = row.get("GBP_pnl", 0.0)
        _fmt_num(ws3, i, 3, pnl_val, pnl=(pnl_val != 0))
        _fmt_num(ws3, i, 4, row.get("GBP_comms", 0.0),    alt=alt)
        cum_val = row.get("GBP_cum", 0.0)
        _fmt_num(ws3, i, 5, cum_val, pnl=(cum_val != 0))

    # Sheet 4: Asset Daily Detail (GBP) ---------------------------------------
    ws4 = wb.create_sheet("Asset Daily Detail (GBP)")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = "A2"
    ws4.sheet_properties.tabColor = _MID_NAVY

    _write_headers(ws4, 1,
        ["Instrument", "Orig CCY", "Date", "Long", "Short",
         "Realized P&L (GBP)", "Total Comms (GBP)", "Net P&L (GBP)", "Cumulative P&L (GBP)"],
        widths=[24, 10, 14, 10, 10, 20, 20, 18, 22])

    row_n = 2
    for d in gbp_daily:
        alt = (row_n % 2 == 0)
        _txt(ws4, row_n, 1, d["instrument"],                  alt=alt)
        _txt(ws4, row_n, 2, d.get("original_currency", ""),   alt=alt, align="center")
        _txt(ws4, row_n, 3, d["date"],                        alt=alt, align="center")
        _txt(ws4, row_n, 4, int(d["long"]),                   alt=alt, align="center")
        _txt(ws4, row_n, 5, int(d["short"]),                  alt=alt, align="center")
        _fmt_num(ws4, row_n, 6, d["realized_pnl"], alt=alt)
        _fmt_num(ws4, row_n, 7, d["total_comms"],  alt=alt)
        _fmt_num(ws4, row_n, 8, d["net_pnl"],      pnl=(d["net_pnl"] != 0))
        _fmt_num(ws4, row_n, 9, d["cumulative"],   pnl=(d["cumulative"] != 0))
        row_n += 1

    # Sheet 5: Commission Analysis (GBP) --------------------------------------
    ws5 = wb.create_sheet("Commission Analysis (GBP)")
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = "A2"
    ws5.sheet_properties.tabColor = _MID_NAVY

    _write_headers(ws5, 1,
        ["Instrument", "Orig CCY", "Gross P&L (GBP)", "Net P&L (GBP)",
         "Broker Comms (GBP)", "Market Fees (GBP)", "NFA Fees (GBP)",
         "Total Comms (GBP)", "Comm Drag %"],
        widths=[24, 10, 18, 18, 18, 18, 14, 18, 14])

    for i, a in enumerate(sorted(gbp_assets, key=lambda x: abs(x["total_comms"]), reverse=True), 2):
        alt = (i % 2 == 0)
        _txt(ws5, i, 1, a["instrument"],                  alt=alt)
        _txt(ws5, i, 2, a.get("original_currency", ""),   alt=alt, align="center")
        _fmt_num(ws5, i, 3, a["realized_pnl"],            alt=alt)
        _fmt_num(ws5, i, 4, a["net_pnl"],                 pnl=True)
        _fmt_num(ws5, i, 5, a["commission_fees"],          alt=alt)
        _fmt_num(ws5, i, 6, a["market_fees"],              alt=alt)
        _fmt_num(ws5, i, 7, a["nfa_fees"],                 alt=alt)
        _fmt_num(ws5, i, 8, a["total_comms"],              alt=alt)
        _txt(ws5, i, 9, f"{a['comm_drag_pct']:.1f}%", alt=alt, align="center")

    tot_row = len(gbp_assets) + 3
    ws5.cell(row=tot_row, column=1, value="TOTAL (GBP)").font = Font(bold=True, name="Calibri", size=10)
    _fmt_num(ws5, tot_row, 5, gbp_cb["commission_fees"], bold=True)
    _fmt_num(ws5, tot_row, 6, gbp_cb["market_fees"],     bold=True)
    _fmt_num(ws5, tot_row, 7, gbp_cb["nfa_fees"],        bold=True)
    _fmt_num(ws5, tot_row, 8, gbp_cb["total"],           bold=True)

    # Sheet 6: FX Rate Log ----------------------------------------------------
    ws6 = wb.create_sheet("FX Rate Log")
    ws6.sheet_view.showGridLines = False
    ws6.freeze_panes = "A2"
    ws6.sheet_properties.tabColor = "2E4F7A"

    fx_cols = [
        "Date", "Instrument", "Orig CCY", "Long", "Short",
        "Realized P&L (Orig)", "Total Comms (Orig)", "Net P&L (Orig)",
        "GBP Rate Used",
        "Realized P&L (GBP)", "Total Comms (GBP)", "Net P&L (GBP)", "Cumulative (GBP)",
    ]
    fx_widths = [14, 24, 10, 10, 10, 20, 20, 18, 16, 20, 20, 18, 20]
    _write_headers(ws6, 1, fx_cols, widths=fx_widths)

    row_n = 2
    for d in gbp_daily:
        alt = (row_n % 2 == 0)
        _txt(ws6, row_n, 1,  d["date"],                        alt=alt, align="center")
        _txt(ws6, row_n, 2,  d["instrument"],                  alt=alt)
        _txt(ws6, row_n, 3,  d.get("original_currency", ""),   alt=alt, align="center")
        _txt(ws6, row_n, 4,  int(d["long"]),                   alt=alt, align="center")
        _txt(ws6, row_n, 5,  int(d["short"]),                  alt=alt, align="center")
        # Native original values
        _fmt_num(ws6, row_n, 6,  d.get("realized_pnl_native", 0), alt=alt)
        _fmt_num(ws6, row_n, 7,  d.get("total_comms_native",  0), alt=alt)
        _fmt_num(ws6, row_n, 8,  d.get("net_pnl_native",      0), pnl=(d.get("net_pnl_native", 0) != 0))
        # Rate used
        rate_cell = ws6.cell(row=row_n, column=9, value=d.get("gbp_rate", 1.0))
        rate_cell.alignment = _right()
        rate_cell.border    = _thin_border()
        rate_cell.number_format = "0.000000"
        rate_cell.fill = _alt_fill() if alt else _white_fill()
        # GBP converted values
        _fmt_num(ws6, row_n, 10, d["realized_pnl"], alt=alt)
        _fmt_num(ws6, row_n, 11, d["total_comms"],  alt=alt)
        _fmt_num(ws6, row_n, 12, d["net_pnl"],      pnl=(d["net_pnl"] != 0))
        _fmt_num(ws6, row_n, 13, d["cumulative"],   pnl=(d["cumulative"] != 0))
        row_n += 1

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()



# --- Excel Export -------------------------------------------------------------

def generate_excel_report(analysis_id: str, trader: str, account: str, gbp: bool = False) -> bytes:
    data = get_analysis(analysis_id)
    if not data:
        raise ValueError("Analysis not found or expired -- please re-upload the file")
    if gbp:
        return _generate_gbp_excel_report(data, trader, account)

    wb = openpyxl.Workbook()

    # -- Sheet 1: Executive Summary --------------------------------------------
    ws1 = wb.active
    ws1.title = 'Executive Summary'
    ws1.sheet_view.showGridLines = False
    ws1.sheet_properties.tabColor = '1F3864'

    ws1.merge_cells('A1:K1')
    t = ws1['A1']
    t.value     = 'AXIA STRATEGY -- TRADE ANALYSIS REPORT'
    t.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=16)
    t.fill      = _hdr_fill()
    t.alignment = _ctr()
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells('A2:K2')
    s = ws1['A2']
    s.value     = (f'Trader: {trader}   |   Account: {account}   |   '
                   f'Period: {data["date_range"]["from"]} -> {data["date_range"]["to"]}   |   '
                   f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}')
    s.font      = Font(italic=True, color='FFFFFF', name='Calibri', size=10)
    s.fill      = PatternFill('solid', fgColor=_MID_NAVY)
    s.alignment = _ctr()
    ws1.row_dimensions[2].height = 20

    ws1.cell(row=4, column=1, value='PERIOD SUMMARY').font = Font(bold=True, name='Calibri', size=11, color=_NAVY)
    period_rows = [
        ('Date From',          data['date_range']['from']),
        ('Date To',            data['date_range']['to']),
        ('Trading Days',       data['date_range']['trading_days']),
        ('Instruments Traded', data['summary']['total_instruments']),
        ('Total Long Lots',    int(data['summary']['total_long_lots'])),
        ('Total Short Lots',   int(data['summary']['total_short_lots'])),
        ('Trade Rows',         data['summary']['total_trades']),
        ('Currencies',         ', '.join(data['summary']['currencies'])),
    ]
    _write_headers(ws1, 5, ['Metric', 'Value'], start_col=1, widths=[22, 18])
    for i, (label, val) in enumerate(period_rows, 6):
        alt = (i % 2 == 0)
        _txt(ws1, i, 1, label, bold=True, alt=alt)
        _txt(ws1, i, 2, val, alt=alt)

    ws1.cell(row=4, column=4, value='P&L BY CURRENCY').font = Font(bold=True, name='Calibri', size=11, color=_NAVY)
    _write_headers(ws1, 5, ['Currency', 'Realized P&L', 'Total Comms', 'Net P&L', 'Instruments'],
                   start_col=4, widths=[12, 16, 14, 16, 14])
    ccy_items = sorted(data['by_currency'].items())
    for i, (ccy, vals) in enumerate(ccy_items, 6):
        alt = (i % 2 == 0)
        count = sum(1 for a in data['by_asset'] if a['currency'] == ccy)
        _txt(ws1, i, 4, ccy, bold=True, alt=alt)
        _fmt_num(ws1, i, 5, vals['realized_pnl'], alt=alt)
        _fmt_num(ws1, i, 6, vals['total_comms'],  alt=alt)
        _fmt_num(ws1, i, 7, vals['net_pnl'],      pnl=True)
        _txt(ws1, i, 8, count, alt=alt)

    ws1.cell(row=4, column=10, value='COMMISSION BREAKDOWN').font = Font(bold=True, name='Calibri', size=11, color=_NAVY)
    _write_headers(ws1, 5, ['Fee Type', 'Amount'], start_col=10, widths=[22, 16])
    cb = data['commission_breakdown']
    comm_rows = [
        ('Broker / Exchange Commission', cb['commission_fees']),
        ('Market / Clearing Fees',       cb['market_fees']),
        ('NFA Regulatory Fees',          cb['nfa_fees']),
        ('TOTAL COMMISSIONS',            cb['total']),
    ]
    for i, (label, val) in enumerate(comm_rows, 6):
        alt = (i % 2 == 0)
        bold = label.startswith('TOTAL')
        _txt(ws1, i, 10, label, bold=bold, alt=alt)
        _fmt_num(ws1, i, 11, val, bold=bold, alt=alt)

    ws1.cell(row=14, column=1, value='* TOP PERFORMERS').font = Font(bold=True, name='Calibri', size=11, color='006100')
    _write_headers(ws1, 15, ['Instrument', 'CCY', 'Long Lots', 'Short Lots', 'Net P&L'], start_col=1)
    for i, a in enumerate(data['top_winners'], 16):
        alt = (i % 2 == 0)
        _txt(ws1, i, 1, a['instrument'], alt=alt)
        _txt(ws1, i, 2, a['currency'],   alt=alt, align='center')
        _txt(ws1, i, 3, int(a['long']),  alt=alt, align='center')
        _txt(ws1, i, 4, int(a['short']), alt=alt, align='center')
        _fmt_num(ws1, i, 5, a['net_pnl'], pnl=True)

    ws1.cell(row=14, column=7, value='v WORST PERFORMERS').font = Font(bold=True, name='Calibri', size=11, color='9C0006')
    _write_headers(ws1, 15, ['Instrument', 'CCY', 'Long Lots', 'Short Lots', 'Net P&L'], start_col=7)
    for i, a in enumerate(data['top_losers'], 16):
        alt = (i % 2 == 0)
        _txt(ws1, i, 7,  a['instrument'], alt=alt)
        _txt(ws1, i, 8,  a['currency'],   alt=alt, align='center')
        _txt(ws1, i, 9,  int(a['long']),  alt=alt, align='center')
        _txt(ws1, i, 10, int(a['short']), alt=alt, align='center')
        _fmt_num(ws1, i, 11, a['net_pnl'], pnl=True)

    ws1.freeze_panes = 'A3'

    # -- Sheet 2: Asset Breakdown ----------------------------------------------
    ws2 = wb.create_sheet('Asset Breakdown')
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = 'A2'
    ws2.sheet_properties.tabColor = _MID_NAVY

    a_cols   = ['Instrument', 'CCY', 'Long Lots', 'Short Lots', 'Trade Days',
                'Realized P&L', 'Broker Comms', 'Market Fees', 'NFA Fees',
                'Total Comms', 'Net P&L', 'Comm Drag %']
    a_widths = [24, 8, 12, 12, 12, 16, 16, 16, 12, 16, 16, 14]
    _write_headers(ws2, 1, a_cols, widths=a_widths)

    for i, a in enumerate(sorted(data['by_asset'], key=lambda x: x['net_pnl']), 2):
        alt = (i % 2 == 0)
        _txt(ws2, i, 1,  a['instrument'],       alt=alt)
        _txt(ws2, i, 2,  a['currency'],         alt=alt, align='center')
        _txt(ws2, i, 3,  int(a['long']),        alt=alt, align='center')
        _txt(ws2, i, 4,  int(a['short']),       alt=alt, align='center')
        _txt(ws2, i, 5,  a['trade_days'],       alt=alt, align='center')
        _fmt_num(ws2, i, 6,  a['realized_pnl'],    alt=alt)
        _fmt_num(ws2, i, 7,  a['commission_fees'],  alt=alt)
        _fmt_num(ws2, i, 8,  a['market_fees'],      alt=alt)
        _fmt_num(ws2, i, 9,  a['nfa_fees'],         alt=alt)
        _fmt_num(ws2, i, 10, a['total_comms'],      alt=alt)
        _fmt_num(ws2, i, 11, a['net_pnl'],          pnl=True)
        _txt(ws2, i, 12, f"{a['comm_drag_pct']:.1f}%", alt=alt, align='center')

    row_n = len(data['by_asset']) + 3
    for ccy, vals in sorted(data['by_currency'].items()):
        ws2.cell(row=row_n, column=1, value=f'{ccy} SUBTOTAL').font = Font(bold=True, name='Calibri', size=10)
        ws2.cell(row=row_n, column=2, value=ccy).font = Font(bold=True, name='Calibri', size=10)
        _fmt_num(ws2, row_n, 11, vals['net_pnl'],     pnl=True)
        _fmt_num(ws2, row_n, 10, vals['total_comms'],  bold=True)
        _fmt_num(ws2, row_n, 6,  vals['realized_pnl'], bold=True)
        row_n += 1

    # -- Sheet 3: Daily P&L ----------------------------------------------------
    ws3 = wb.create_sheet('Daily P&L')
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = 'A2'
    ws3.sheet_properties.tabColor = _MID_NAVY

    ccys = data['summary']['currencies']
    d_cols   = ['Date', 'Total Lots'] + [f'{c} P&L' for c in ccys] + [f'{c} Comms' for c in ccys]
    d_widths = [14, 12] + [14] * len(ccys) + [14] * len(ccys)
    _write_headers(ws3, 1, d_cols, widths=d_widths)

    for i, row in enumerate(data['by_date'], 2):
        alt = (i % 2 == 0)
        _txt(ws3, i, 1, row['date'],                       alt=alt, align='center')
        _txt(ws3, i, 2, int(row.get('total_lots', 0)),     alt=alt, align='center')
        for j, c in enumerate(ccys, 3):
            pnl_val = row.get(f'{c}_pnl', 0.0)
            _fmt_num(ws3, i, j, pnl_val, pnl=(pnl_val != 0))
        for j, c in enumerate(ccys, 3 + len(ccys)):
            _fmt_num(ws3, i, j, row.get(f'{c}_comms', 0.0), alt=alt)

    # -- Sheet 4: Asset Daily Detail -------------------------------------------
    ws4 = wb.create_sheet('Asset Daily Detail')
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = 'A2'
    ws4.sheet_properties.tabColor = _MID_NAVY

    dd_cols   = ['Instrument', 'CCY', 'Date', 'Long', 'Short',
                 'Realized P&L', 'Total Comms', 'Net P&L', 'Cumulative P&L']
    dd_widths = [24, 8, 14, 10, 10, 16, 14, 16, 18]
    _write_headers(ws4, 1, dd_cols, widths=dd_widths)

    row_n = 2
    for d in data.get('portfolio_daily_detail', []):
        alt = (row_n % 2 == 0)
        _txt(ws4, row_n, 1, d['instrument'],      alt=alt)
        _txt(ws4, row_n, 2, d['currency'],        alt=alt, align='center')
        _txt(ws4, row_n, 3, d['date'],            alt=alt, align='center')
        _txt(ws4, row_n, 4, int(d['long']),       alt=alt, align='center')
        _txt(ws4, row_n, 5, int(d['short']),      alt=alt, align='center')
        _fmt_num(ws4, row_n, 6, d['realized_pnl'], alt=alt)
        _fmt_num(ws4, row_n, 7, d['total_comms'],  alt=alt)
        _fmt_num(ws4, row_n, 8, d['net_pnl'],      pnl=(d['net_pnl'] != 0))
        _fmt_num(ws4, row_n, 9, d['cumulative'],   pnl=(d['cumulative'] != 0))
        row_n += 1

    # -- Sheet 5: Commission Analysis ------------------------------------------
    ws5 = wb.create_sheet('Commission Analysis')
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = 'A2'
    ws5.sheet_properties.tabColor = _MID_NAVY

    ca_cols   = ['Instrument', 'CCY', 'Gross P&L', 'Net P&L',
                 'Broker Comms', 'Market Fees', 'NFA Fees', 'Total Comms',
                 'Comms as % of Gross', 'Comm Drag on Net']
    ca_widths = [24, 8, 16, 16, 14, 14, 12, 14, 20, 20]
    _write_headers(ws5, 1, ca_cols, widths=ca_widths)

    for i, a in enumerate(sorted(data['by_asset'], key=lambda x: abs(x['total_comms']), reverse=True), 2):
        alt = (i % 2 == 0)
        _txt(ws5, i, 1, a['instrument'], alt=alt)
        _txt(ws5, i, 2, a['currency'],   alt=alt, align='center')
        _fmt_num(ws5, i, 3,  a['realized_pnl'],            alt=alt)
        _fmt_num(ws5, i, 4,  a['net_pnl'],                 pnl=True)
        _fmt_num(ws5, i, 5,  a['commission_fees'],          alt=alt)
        _fmt_num(ws5, i, 6,  a['market_fees'],              alt=alt)
        _fmt_num(ws5, i, 7,  a['nfa_fees'],                 alt=alt)
        _fmt_num(ws5, i, 8,  a['total_comms'],              alt=alt)
        _txt(ws5, i, 9,  f"{a['comm_drag_pct']:.1f}%", alt=alt, align='center')
        gross = a['realized_pnl']
        drag  = round((abs(a['total_comms']) / abs(gross) * 100), 1) if gross != 0 else 0
        _txt(ws5, i, 10, f"{drag:.1f}%", alt=alt, align='center')

    tot_row = len(data['by_asset']) + 3
    ws5.cell(row=tot_row, column=1, value='TOTAL').font = Font(bold=True, name='Calibri', size=10)
    cb = data['commission_breakdown']
    _fmt_num(ws5, tot_row, 5, cb['commission_fees'], bold=True)
    _fmt_num(ws5, tot_row, 6, cb['market_fees'],     bold=True)
    _fmt_num(ws5, tot_row, 7, cb['nfa_fees'],        bold=True)
    _fmt_num(ws5, tot_row, 8, cb['total'],           bold=True)

    # -- Save ------------------------------------------------------------------
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
