# backend/src/services/merge_service.py
"""
Merge multiple AXIA statement Excel files into a single chronological workbook.

Reads data rows from each .xlsx (skips headers and TOTALS rows),
sorts by TRADE DATE ascending, outputs via write_statement_excel.
"""

from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import List

import openpyxl

from src.services.excel_writer import write_statement_excel

# Column index mapping (1-based) matching excel_writer.COLUMNS order
_COL = {
    "TRADE DATE":         1,
    "DELIVERY / PRODUCT": 2,
    "LONG":               3,
    "SHORT":              4,
    "REALIZED PnL":       5,
    "COMMISION FEES":     6,
    "MARKET FEES":        7,
    "NFA FEES":           8,
    "TOTAL COMMS":        9,
    "TOTAL PnL":          10,
    "CURRENCY":           11,
}


def _read_excel_rows(path: Path) -> list[dict]:
    """Read data rows from a single AXIA statement Excel file."""
    wb = openpyxl.load_workbook(str(path), data_only=True)

    # Use the first sheet (Statement)
    ws = None
    for name in wb.sheetnames:
        if name.lower() != "info":
            ws = wb[name]
            break
    if ws is None:
        return []

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not any(row):
            continue
        # Skip TOTALS rows (col A = "TOTALS" string)
        first = row[0]
        if isinstance(first, str) and first.strip().upper() in ("TOTALS", "TOTAL"):
            continue

        trade_date = row[_COL["TRADE DATE"] - 1]
        # Must have a valid date — skip footer/summary rows
        if trade_date is None:
            continue
        if isinstance(trade_date, str):
            # Try parse if stored as string
            try:
                trade_date = datetime.strptime(trade_date, "%Y-%m-%d")
            except ValueError:
                try:
                    trade_date = datetime.strptime(trade_date, "%d-%b-%Y")
                except ValueError:
                    continue  # unparseable → skip

        td_str = trade_date.strftime("%Y-%m-%d") if isinstance(trade_date, datetime) else str(trade_date)

        def _v(col_name):
            val = row[_COL[col_name] - 1]
            # TOTAL COMMS / TOTAL PnL columns may be formula results
            return val

        rows.append({
            "trade_date":       td_str,
            "delivery_product": row[_COL["DELIVERY / PRODUCT"] - 1] or "",
            "long":             row[_COL["LONG"] - 1],
            "short":            row[_COL["SHORT"] - 1],
            "realized_pnl":     row[_COL["REALIZED PnL"] - 1],
            "commission_fees":  row[_COL["COMMISION FEES"] - 1],
            "market_fees":      row[_COL["MARKET FEES"] - 1],
            "nfa_fees":         row[_COL["NFA FEES"] - 1],
            "currency":         row[_COL["CURRENCY"] - 1],
        })
    return rows


def merge_statements(excel_paths: List[Path]) -> bytes:
    """
    Merge multiple AXIA statement Excel files.
    Returns merged Excel bytes sorted by trade_date ASC.
    """
    all_rows = []
    for path in excel_paths:
        all_rows.extend(_read_excel_rows(path))

    if not all_rows:
        raise ValueError("No data rows found in provided Excel files.")

    # Sort by trade_date ascending
    all_rows.sort(key=lambda r: r["trade_date"])

    # Derive date range for merged filename hint
    dates      = [r["trade_date"] for r in all_rows]
    date_from  = min(dates)
    date_to    = max(dates)

    merged_data = {
        "trade_date": f"{date_from} to {date_to}",
        "client":     None,
        "account":    None,
        "rows":       all_rows,
    }

    return write_statement_excel(merged_data)
